import pandas as pd
from collections import defaultdict
import os
import json
import jieba
from concurrent.futures import ProcessPoolExecutor
from itertools import chain
import PyPDF2
import docx
from openpyxl import load_workbook


# 支持的文件格式
SUPPORTED_FILE_FORMATS = [".pdf", ".doc", ".docx", ".xls", ".xlsx"]


# 加载停用词
def load_stopwords(stopword_files):
    """加载停用词表"""
    stopwords = set()
    for file in stopword_files:
        with open(file, "r", encoding="utf-8") as f:
            stopwords.update(line.strip() for line in f)
    return stopwords


# 停用词表路径
STOPWORDS_FILES = ["cn_stopwords.txt", "baidu_stopwords.txt"]
STOPWORDS = load_stopwords(STOPWORDS_FILES)


# 加载数据
def load_raw_data(file_path, chunk_size=10000):
    """按块加载数据，减少内存占用"""
    chunks = pd.read_csv(file_path, chunksize=chunk_size)
    return chunks


# 判断是否为文件链接
def is_file_link(url):
    for ext in SUPPORTED_FILE_FORMATS:
        if url.lower().endswith(ext):
            return True
    return False


# 提取文件内容
def extract_file_content(file_path):
    """根据文件类型提取文字内容"""
    if not os.path.exists(file_path):
        return "[FILE NOT FOUND]"

    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    try:
        if ext == ".pdf":
            return extract_pdf_content(file_path)
        elif ext in [".doc", ".docx"]:
            return extract_doc_content(file_path)
        elif ext in [".xls", ".xlsx"]:
            return extract_xls_content(file_path)
        else:
            return "[UNSUPPORTED FILE FORMAT]"
    except Exception as e:
        return f"[ERROR READING FILE: {e}]"


def extract_pdf_content(file_path):
    """提取 PDF 文件内容"""
    try:
        with open(file_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            content = []
            for page in reader.pages:
                content.append(page.extract_text())
            return " ".join(content)
    except Exception:
        return "[UNREADABLE PDF]"


def extract_doc_content(file_path):
    """提取 Word 文件内容"""
    try:
        doc = docx.Document(file_path)
        content = [p.text for p in doc.paragraphs]
        return " ".join(content)
    except Exception:
        return "[UNREADABLE WORD FILE]"


def extract_xls_content(file_path):
    """提取 Excel 文件内容"""
    try:
        wb = load_workbook(file_path, data_only=True)
        content = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                content.extend(str(cell) for cell in row if cell is not None)
        return " ".join(content)
    except Exception:
        return "[UNREADABLE EXCEL FILE]"


# 构建倒排索引
def build_inverted_index(rows):
    """构建部分倒排索引"""
    inverted_index = defaultdict(list)

    for row in rows:
        doc_id = row["url"]
        content = ""
        doc_number = row["line_number"]

        # 处理全文
        if is_file_link(doc_id):
            # 如果是文件链接，尝试提取文件内容
            content = extract_file_content(row["body"])
            content = row["url"] + content
        else:
            # 拼接 title, anchor_texts, 和 body 内容
            content = f"{row['title']}{row['url']}{row['anchor_texts']} {row['body']}"

        # 处理标题
        # if is_file_link(doc_id):
        #     content = extract_file_content(row["body"])
        #     content = row["title"] + row["url"] + content
        # else:
        #     # 拼接 title, anchor_texts, 和 body 内容
        #     content = f"{row['title']}{row['url']}"

        # 处理文件
        # if is_file_link(doc_id):
        #     content = extract_file_content(row["body"])
        #     content = row["url"] + content
        # else:
        #     continue
        # 分词并索引
        words = jieba.lcut(content)  # 使用 jieba 分词
        for position, word in enumerate(words):
            word = word.lower().strip(",.!?;:\"'()[]{}")  # 标准化词
            if word and word not in STOPWORDS:  # 过滤停用词
                inverted_index[word].append(
                    (str(doc_number), position)  # 将 doc_number 转换为字符串
                )

    return inverted_index


# 合并多个倒排索引
def merge_inverted_indexes(indexes):
    """合并多个部分倒排索引"""
    merged_index = defaultdict(list)
    for index in indexes:
        for word, postings in index.items():
            merged_index[word].extend(postings)
    return merged_index


# 保存分块倒排索引
def save_inverted_index_in_chunks(inverted_index, output_dir="inverted_index_chunks"):
    """分块保存倒排索引"""
    os.makedirs(output_dir, exist_ok=True)
    chunk_files = defaultdict(dict)

    # 分块存储
    for word, postings in inverted_index.items():
        if "\u4e00" <= word[0] <= "\u9fff":  # 中文字符范围
            chunk_key = f"chinese_{word[0]}"
        elif word[0].isalpha():  # 英文字母
            chunk_key = f"alpha_{word[0].lower()}"
        elif word[0].isdigit():  # 数字
            chunk_key = "numeric"
        else:  # 其他符号统一存放
            chunk_key = "others"

        chunk_files[chunk_key][word] = postings

    # 保存到对应的 JSON 文件
    for chunk_key, chunk_data in chunk_files.items():
        chunk_file = os.path.join(output_dir, f"{chunk_key}.json")
        try:
            with open(chunk_file, "a", encoding="utf-8") as f:
                json.dump(chunk_data, f, ensure_ascii=False, indent=4)
            print(f"分块保存成功: {chunk_file}")
        except Exception as e:
            print(f"[ERROR SAVING CHUNK] 文件: {chunk_file}, 错误: {e}")
            # 可以选择记录到日志文件或其他地方，方便后续检查问题
            with open(
                os.path.join(output_dir, "error_log.txt"), "a", encoding="utf-8"
            ) as log_file:
                log_file.write(f"保存失败的分块: {chunk_key}\n")
                log_file.write(f"错误信息: {e}\n")
                log_file.write("\n")


# 并行构建倒排索引
def parallel_build_inverted_index(file_path, max_workers=4, chunk_size=10000):
    """并行构建倒排索引"""
    chunks = load_raw_data(file_path, chunk_size=chunk_size)
    partial_indexes = []

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = [
            executor.submit(build_inverted_index, chunk.to_dict("records"))
            for chunk in chunks
        ]
        for future in futures:
            partial_indexes.append(future.result())

    # 合并结果
    final_index = merge_inverted_indexes(partial_indexes)
    return final_index


# # 主函数
def main():
    file_path = "../crawler/linenumber_title_url_anchor_body.csv"  # 数据文件路径
    max_workers = 4  # 并行线程数
    chunk_size = 10000  # 每块大小

    # 并行构建倒排索引
    print("开始构建倒排索引...")
    inverted_index = parallel_build_inverted_index(
        file_path, max_workers=max_workers, chunk_size=chunk_size
    )

    # 分块保存倒排索引
    save_inverted_index_in_chunks(inverted_index)
    print("倒排索引已分块保存到 'inverted_index_chunks' 目录！")


if __name__ == "__main__":
    main()
