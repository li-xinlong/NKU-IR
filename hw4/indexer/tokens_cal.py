import pandas as pd
import os
import json
import jieba
from concurrent.futures import ProcessPoolExecutor
from collections import defaultdict
import PyPDF2
import docx
from openpyxl import load_workbook

# 支持的文件格式
SUPPORTED_FILE_FORMATS = [".pdf", ".doc", ".docx", ".xls", ".xlsx"]


# 加载停用词
def load_stopwords(stopword_files):
    """加载停用词表"""
    stopwords = set()
    for file in stopword_files:
        with open(file, "r", encoding="utf-8") as f:
            stopwords.update(line.strip() for line in f)
    return stopwords


# 停用词表路径
STOPWORDS_FILES = ["cn_stopwords.txt", "baidu_stopwords.txt"]
STOPWORDS = load_stopwords(STOPWORDS_FILES)


# 加载数据
def load_raw_data(file_path, chunk_size=10000):
    """按块加载数据，减少内存占用"""
    chunks = pd.read_csv(file_path, chunksize=chunk_size)
    return chunks


# 判断是否为文件链接
def is_file_link(url):
    for ext in SUPPORTED_FILE_FORMATS:
        if url.lower().endswith(ext):
            return True
    return False


# 提取文件内容
def extract_file_content(file_path):
    """根据文件类型提取文字内容"""
    if not os.path.exists(file_path):
        return "[FILE NOT FOUND]"

    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    try:
        if ext == ".pdf":
            return extract_pdf_content(file_path)
        elif ext in [".doc", ".docx"]:
            return extract_doc_content(file_path)
        elif ext in [".xls", ".xlsx"]:
            return extract_xls_content(file_path)
        else:
            return "[UNSUPPORTED FILE FORMAT]"
    except Exception as e:
        return f"[ERROR READING FILE: {e}]"


def extract_pdf_content(file_path):
    """提取 PDF 文件内容"""
    try:
        with open(file_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            content = []
            for page in reader.pages:
                content.append(page.extract_text())
            return " ".join(content)
    except Exception:
        return "[UNREADABLE PDF]"


def extract_doc_content(file_path):
    """提取 Word 文件内容"""
    try:
        doc = docx.Document(file_path)
        content = [p.text for p in doc.paragraphs]
        return " ".join(content)
    except Exception:
        return "[UNREADABLE WORD FILE]"


def extract_xls_content(file_path):
    """提取 Excel 文件内容"""
    try:
        wb = load_workbook(file_path, data_only=True)
        content = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                content.extend(str(cell) for cell in row if cell is not None)
        return " ".join(content)
    except Exception:
        return "[UNREADABLE EXCEL FILE]"


# 统计每个文档中的单词数
def count_words_in_document(rows):
    """统计每个文档的单词数"""
    word_count_data = []

    for row in rows:
        doc_id = row["url"]
        content = ""
        doc_number = row["line_number"]

        # 处理全文
        if is_file_link(doc_id):
            # 如果是文件链接，尝试提取文件内容
            content = extract_file_content(row["body"])
            content = row["url"] + content
        else:
            # 拼接 title, anchor_texts, 和 body 内容
            content = f"{row['title']}{row['url']}{row['anchor_texts']} {row['body']}"

        # 处理标题
        # if is_file_link(doc_id):
        #     content = extract_file_content(row["body"])
        #     content = row["title"] + row["url"] + content
        # else:
        #     # 拼接 title, anchor_texts, 和 body 内容
        #     content = f"{row['title']}{row['url']}"

        # 处理文件
        # if is_file_link(doc_id):
        #     content = extract_file_content(row["body"])
        #     content = row["url"] + content
        # else:
        #     continue

        # 分词并清理
        words = jieba.lcut(content)  # 使用 jieba 分词
        filtered_words = [
            word.lower().strip(",.!?;:\"'()[]{}")
            for word in words
            if word and word not in STOPWORDS
        ]

        # 记录文档的单词数
        word_count_data.append(
            {"linenumber": doc_number, "url": doc_id, "word_count": len(filtered_words)}
        )

    return word_count_data


# 合并多个文档的单词数统计
def merge_word_count_results(results):
    """合并多个结果"""
    merged_data = []
    for result in results:
        merged_data.extend(result)
    return merged_data


# 保存统计的单词数
def save_word_count(word_count_data, output_file="title_word_count.csv"):
    """将单词统计结果保存到 CSV 文件"""
    df = pd.DataFrame(word_count_data)
    df.to_csv(output_file, index=False, columns=["linenumber", "url", "word_count"])
    print(f"单词数统计结果已保存到 {output_file}")


# 并行统计文档中的单词数
def parallel_count_words(file_path, max_workers=4, chunk_size=10000):
    """并行统计文档中的单词数"""
    chunks = load_raw_data(file_path, chunk_size=chunk_size)
    partial_results = []

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = [
            executor.submit(count_words_in_document, chunk.to_dict("records"))
            for chunk in chunks
        ]
        for future in futures:
            partial_results.append(future.result())

    # 合并结果
    final_results = merge_word_count_results(partial_results)
    return final_results


# 主函数
def main():
    file_path = "../crawler/linenumber_title_url_anchor_body.csv"  # 数据文件路径
    max_workers = 4  # 并行线程数
    chunk_size = 10000  # 每块大小

    # 并行统计文档中的单词数
    print("开始统计文档中的单词数...")
    word_count_data = parallel_count_words(
        file_path, max_workers=max_workers, chunk_size=chunk_size
    )

    # 保存统计结果
    save_word_count(word_count_data)
    print("单词数统计完成！")


if __name__ == "__main__":
    main()
